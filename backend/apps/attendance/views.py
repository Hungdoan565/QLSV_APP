from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
import uuid
import qrcode
import io
import base64
from datetime import datetime, timedelta
from .models import Attendance, AttendanceSession
from .serializers import AttendanceSerializer, AttendanceSessionSerializer


class AttendanceListCreateView(generics.ListCreateAPIView):
    """List and create attendance records"""
    queryset = Attendance.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AttendanceSerializer
    
    def get_queryset(self):
        queryset = Attendance.objects.all()
        session_id = self.request.query_params.get('session_id', None)
        student_id = self.request.query_params.get('student_id', None)
        
        if session_id is not None:
            queryset = queryset.filter(session_id=session_id)
        if student_id is not None:
            queryset = queryset.filter(student_id=student_id)
            
        return queryset.order_by('-created_at')


class AttendanceDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete an attendance record"""
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]


class AttendanceSessionListCreateView(generics.ListCreateAPIView):
    """List and create attendance sessions"""
    queryset = AttendanceSession.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AttendanceSessionSerializer
    
    def get_queryset(self):
        queryset = AttendanceSession.objects.all()
        class_id = self.request.query_params.get('class_id', None)
        
        if class_id is not None:
            queryset = queryset.filter(class_instance_id=class_id)
            
        return queryset.order_by('-session_date', '-start_time')


class AttendanceSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete an attendance session"""
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def attendance_statistics(request):
    """Get attendance statistics"""
    total_sessions = AttendanceSession.objects.count()
    total_attendances = Attendance.objects.count()
    present_count = Attendance.objects.filter(status='present').count()
    
    return Response({
        'total_sessions': total_sessions,
        'total_attendances': total_attendances,
        'present_count': present_count,
        'attendance_rate': round((present_count / total_attendances * 100), 2) if total_attendances > 0 else 0,
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_qr_code(request, session_id):
    """Generate QR code for attendance session"""
    try:
        session = AttendanceSession.objects.get(id=session_id)
        
        # Check permission
        if request.user.role != 'admin' and session.class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền tạo QR code cho buổi điểm danh này'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Generate unique QR code
        qr_code = str(uuid.uuid4())
        session.qr_code = qr_code
        session.save()
        
        # Create QR code image
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_code)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return Response({
            'qr_code': qr_code,
            'qr_image': f'data:image/png;base64,{img_str}',
            'session_info': {
                'id': session.id,
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date,
                'start_time': session.start_time,
                'end_time': session.end_time
            },
            'expires_at': session.end_time
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def check_in_with_qr(request):
    """Check in using QR code"""
    try:
        qr_code = request.data.get('qr_code')
        student_id = request.data.get('student_id')
        
        if not qr_code or not student_id:
            return Response(
                {'error': 'QR code và mã sinh viên là bắt buộc'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Find session by QR code
        session = AttendanceSession.objects.get(qr_code=qr_code, is_active=True)
        
        # Check if session is still active (within time range)
        now = timezone.now()
        session_datetime = timezone.make_aware(
            datetime.combine(session.session_date, session.end_time)
        )
        
        if now > session_datetime:
            return Response(
                {'error': 'Buổi điểm danh đã kết thúc'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Find student
        from apps.students.models import Student
        student = Student.objects.get(student_id=student_id)
        
        # Check if student is in the class
        from apps.classes.models import ClassStudent
        if not ClassStudent.objects.filter(
            class_obj=session.class_obj,
            student=student,
            is_active=True
        ).exists():
            return Response(
                {'error': 'Sinh viên không thuộc lớp này'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if already attended
        attendance, created = Attendance.objects.get_or_create(
            session=session,
            student=student,
            defaults={
                'status': 'present',
                'check_in_time': now
            }
        )
        
        if not created:
            if attendance.status == 'present':
                return Response(
                    {'error': 'Bạn đã điểm danh rồi'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                # Update existing attendance
                attendance.status = 'present'
                attendance.check_in_time = now
                attendance.save()
        
        return Response({
            'message': 'Điểm danh thành công',
            'attendance': AttendanceSerializer(attendance).data,
            'session_info': {
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date
            }
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'QR code không hợp lệ hoặc buổi điểm danh không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
    except Student.DoesNotExist:
        return Response({'error': 'Không tìm thấy sinh viên'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def attendance_analytics(request, session_id):
    """Get attendance analytics for a session"""
    try:
        session = AttendanceSession.objects.get(id=session_id)
        
        # Check permission
        if request.user.role != 'admin' and session.class_obj.teacher != request.user:
            return Response(
                {'error': 'Bạn không có quyền xem thống kê buổi điểm danh này'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get attendance data
        total_students = session.class_obj.current_students_count
        present_count = Attendance.objects.filter(session=session, status='present').count()
        absent_count = total_students - present_count
        attendance_rate = (present_count / total_students * 100) if total_students > 0 else 0
        
        # Get attendance by time
        attendances = Attendance.objects.filter(session=session, status='present')
        attendance_times = []
        for att in attendances:
            attendance_times.append({
                'student_id': att.student.student_id,
                'student_name': att.student.full_name,
                'check_in_time': att.check_in_time,
                'status': att.status
            })
        
        return Response({
            'session_info': {
                'id': session.id,
                'session_name': session.session_name,
                'class_name': session.class_obj.class_name,
                'session_date': session.session_date,
                'start_time': session.start_time,
                'end_time': session.end_time
            },
            'statistics': {
                'total_students': total_students,
                'present_count': present_count,
                'absent_count': absent_count,
                'attendance_rate': round(attendance_rate, 2)
            },
            'attendance_details': attendance_times
        })
        
    except AttendanceSession.DoesNotExist:
        return Response({'error': 'Không tìm thấy buổi điểm danh'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel(request):
    """Import attendance records from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Only Excel files (.xlsx, .xls) are supported'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse Excel file with openpyxl
        from openpyxl import load_workbook
        import io
        
        # Load workbook from uploaded file
        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
        
        # Expected columns mapping
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id'],
            'session_id': ['session_id', 'buoi_diem_danh', 'session'],
            'status': ['status', 'trang_thai', 'attendance_status'],
            'check_in_time': ['check_in_time', 'gio_vao', 'time_in'],
            'check_out_time': ['check_out_time', 'gio_ra', 'time_out'],
            'notes': ['notes', 'ghi_chu', 'note']
        }
        
        # Map headers to expected fields
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            for header in headers:
                if header in possible_names:
                    field_mapping[field] = headers.index(header)
                    break
        
        created_attendance = []
        errors = []
        
        # Process each row
        for row_num in range(2, worksheet.max_row + 1):
            try:
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    row_data[field] = str(cell_value).strip() if cell_value else ''
                
                # Set defaults for missing fields
                if not row_data.get('status'):
                    row_data['status'] = 'present'
                if not row_data.get('check_in_time'):
                    row_data['check_in_time'] = timezone.now().isoformat()
                
                # Validate required fields
                if not row_data.get('student_id') or not row_data.get('session_id'):
                    errors.append({
                        'row': row_num,
                        'error': 'Missing required fields: student_id and session_id',
                        'data': row_data
                    })
                    continue
                
                # Convert session_id to int
                try:
                    row_data['session_id'] = int(row_data['session_id'])
                except ValueError:
                    errors.append({
                        'row': row_num,
                        'error': f'Invalid session_id format: {row_data["session_id"]}',
                        'data': row_data
                    })
                    continue
                
                # Validate status
                valid_statuses = ['present', 'absent', 'late', 'excused']
                if row_data['status'] not in valid_statuses:
                    errors.append({
                        'row': row_num,
                        'error': f'Invalid status: {row_data["status"]}. Must be one of: {valid_statuses}',
                        'data': row_data
                    })
                    continue
                
                # Create attendance record
                serializer = AttendanceSerializer(data=row_data)
                if serializer.is_valid():
                    attendance = serializer.save()
                    created_attendance.append(AttendanceSerializer(attendance).data)
                else:
                    errors.append({
                        'row': row_num,
                        'errors': serializer.errors,
                        'data': row_data
                    })
                    
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': f'Row {row_num} processing failed'
                })
        
        return Response({
            'success': True,
            'message': f'Successfully imported {len(created_attendance)} attendance records from Excel file',
            'created_count': len(created_attendance),
            'created_attendance': created_attendance,
            'errors': errors,
            'details': {
                'total_rows_processed': worksheet.max_row - 1,
                'successful_imports': len(created_attendance),
                'failed_imports': len(errors)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


# QR Code API Endpoints
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_qr_code(request):
    """
    Generate QR code for attendance session
    POST /api/attendance/generate-qr/
    """
    try:
        from .qr_service import QRCodeService
        
        session_id = request.data.get('session_id')
        session_name = request.data.get('session_name', '')
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied. Only the class teacher can generate QR codes.'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Generate QR code
        qr_code_base64, token = QRCodeService.generate_qr_code(
            session_id=session_id,
            teacher_id=request.user.id,
            session_name=session_name or session.session_name
        )
        
        return Response({
            'success': True,
            'qr_code': qr_code_base64,
            'token': token,
            'session_id': session_id,
            'expires_in_minutes': QRCodeService.QR_EXPIRY_MINUTES,
            'session_info': {
                'name': session.session_name,
                'class_name': session.class_obj.class_name,
                'start_time': session.start_time,
                'end_time': session.end_time,
                'location': session.location
            }
        })
        
    except ValueError as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to generate QR code: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def validate_qr_code(request):
    """
    Validate QR code and check in student
    POST /api/attendance/validate-qr/
    """
    try:
        from .qr_service import QRCodeService, QRCodeValidator
        
        qr_token = request.data.get('qr_token')
        student_id = request.data.get('student_id')
        
        if not qr_token or not student_id:
            return Response({
                'success': False,
                'error': 'QR token and student ID are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate QR token
        try:
            validation_result = QRCodeService.validate_qr_token(qr_token)
            session_id = validation_result['session_id']
        except ValueError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate session access
        is_valid_access, access_message = QRCodeValidator.validate_session_access(
            session_id, student_id
        )
        
        if not is_valid_access:
            return Response({
                'success': False,
                'error': access_message
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check for duplicate check-in
        is_duplicate, duplicate_message = QRCodeValidator.check_duplicate_checkin(
            session_id, student_id
        )
        
        if is_duplicate:
            return Response({
                'success': False,
                'error': duplicate_message
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create or update attendance record
        try:
            attendance, created = Attendance.objects.get_or_create(
                session_id=session_id,
                student_id=student_id,
                defaults={
                    'status': 'present',
                    'check_in_time': timezone.now(),
                    'is_late': False,
                    'notes': 'Checked in via QR code'
                }
            )
            
            if not created:
                # Update existing record
                attendance.status = 'present'
                attendance.check_in_time = timezone.now()
                attendance.is_late = False
                attendance.notes = 'Checked in via QR code'
                attendance.save()
            
            # Get session info for response
            session = AttendanceSession.objects.select_related('class_obj').get(id=session_id)
            
            return Response({
                'success': True,
                'message': 'Successfully checked in',
                'attendance_id': attendance.id,
                'session_info': {
                    'id': session.id,
                    'name': session.session_name,
                    'class_name': session.class_obj.class_name,
                    'start_time': session.start_time,
                    'end_time': session.end_time,
                    'location': session.location
                },
                'check_in_time': attendance.check_in_time,
                'is_late': attendance.is_late
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Failed to create attendance record: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'QR validation failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_qr_status(request, session_id):
    """
    Get QR code status for a session
    GET /api/attendance/qr-status/{session_id}/
    """
    try:
        from .qr_service import QRCodeService
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get QR status
        qr_status = QRCodeService.get_qr_status(session_id, request.user.id)
        
        return Response({
            'success': True,
            'qr_status': qr_status,
            'session_info': {
                'id': session.id,
                'name': session.session_name,
                'class_name': session.class_obj.class_name,
                'is_active': session.is_active,
                'start_time': session.start_time,
                'end_time': session.end_time
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to get QR status: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def revoke_qr_code(request, session_id):
    """
    Revoke QR code for a session
    POST /api/attendance/revoke-qr/{session_id}/
    """
    try:
        from .qr_service import QRCodeService
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Revoke QR code
        success = QRCodeService.revoke_qr_token(session_id, request.user.id)
        
        if success:
            return Response({
                'success': True,
                'message': 'QR code revoked successfully'
            })
        else:
            return Response({
                'success': False,
                'error': 'Failed to revoke QR code'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to revoke QR code: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_attendance(request):
    """Export attendance records to Excel file"""
    try:
        import pandas as pd
        from io import BytesIO
        from django.http import HttpResponse
        
        # Get query parameters
        session_id = request.query_params.get('session_id', None)
        class_id = request.query_params.get('class_id', None)
        student_id = request.query_params.get('student_id', None)
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)
        
        # Filter attendance records
        queryset = Attendance.objects.select_related('student', 'session', 'session__class_obj').all()
        
        if session_id:
            queryset = queryset.filter(session_id=session_id)
        if class_id:
            queryset = queryset.filter(session__class_obj_id=class_id)
        if student_id:
            queryset = queryset.filter(student__student_id=student_id)
        if date_from:
            queryset = queryset.filter(session__session_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(session__session_date__lte=date_to)
        
        # Convert to list of dictionaries
        attendance_data = []
        for attendance in queryset:
            attendance_data.append({
                'Student ID': attendance.student.student_id,
                'Student Name': f"{attendance.student.first_name} {attendance.student.last_name}",
                'Class': attendance.session.class_obj.class_name,
                'Session Name': attendance.session.session_name,
                'Session Date': attendance.session.session_date.strftime('%Y-%m-%d'),
                'Start Time': attendance.session.start_time.strftime('%H:%M'),
                'End Time': attendance.session.end_time.strftime('%H:%M'),
                'Location': attendance.session.location or '',
                'Status': attendance.get_status_display(),
                'Check In Time': attendance.check_in_time.strftime('%Y-%m-%d %H:%M') if attendance.check_in_time else '',
                'Check Out Time': attendance.check_out_time.strftime('%Y-%m-%d %H:%M') if attendance.check_out_time else '',
                'Is Late': 'Yes' if attendance.is_late else 'No',
                'Notes': attendance.notes or '',
                'Created Date': attendance.created_at.strftime('%Y-%m-%d %H:%M')
            })
        
        # Create DataFrame
        df = pd.DataFrame(attendance_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Attendance']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename
        filename_parts = ['attendance']
        if session_id:
            filename_parts.append(f'session_{session_id}')
        if class_id:
            filename_parts.append(f'class_{class_id}')
        if student_id:
            filename_parts.append(f'student_{student_id}')
        
        filename = '_'.join(filename_parts) + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


# QR Code API Endpoints
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_qr_code(request):
    """
    Generate QR code for attendance session
    POST /api/attendance/generate-qr/
    """
    try:
        from .qr_service import QRCodeService
        
        session_id = request.data.get('session_id')
        session_name = request.data.get('session_name', '')
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied. Only the class teacher can generate QR codes.'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Generate QR code
        qr_code_base64, token = QRCodeService.generate_qr_code(
            session_id=session_id,
            teacher_id=request.user.id,
            session_name=session_name or session.session_name
        )
        
        return Response({
            'success': True,
            'qr_code': qr_code_base64,
            'token': token,
            'session_id': session_id,
            'expires_in_minutes': QRCodeService.QR_EXPIRY_MINUTES,
            'session_info': {
                'name': session.session_name,
                'class_name': session.class_obj.class_name,
                'start_time': session.start_time,
                'end_time': session.end_time,
                'location': session.location
            }
        })
        
    except ValueError as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to generate QR code: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def validate_qr_code(request):
    """
    Validate QR code and check in student
    POST /api/attendance/validate-qr/
    """
    try:
        from .qr_service import QRCodeService, QRCodeValidator
        
        qr_token = request.data.get('qr_token')
        student_id = request.data.get('student_id')
        
        if not qr_token or not student_id:
            return Response({
                'success': False,
                'error': 'QR token and student ID are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate QR token
        try:
            validation_result = QRCodeService.validate_qr_token(qr_token)
            session_id = validation_result['session_id']
        except ValueError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate session access
        is_valid_access, access_message = QRCodeValidator.validate_session_access(
            session_id, student_id
        )
        
        if not is_valid_access:
            return Response({
                'success': False,
                'error': access_message
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check for duplicate check-in
        is_duplicate, duplicate_message = QRCodeValidator.check_duplicate_checkin(
            session_id, student_id
        )
        
        if is_duplicate:
            return Response({
                'success': False,
                'error': duplicate_message
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create or update attendance record
        try:
            attendance, created = Attendance.objects.get_or_create(
                session_id=session_id,
                student_id=student_id,
                defaults={
                    'status': 'present',
                    'check_in_time': timezone.now(),
                    'is_late': False,
                    'notes': 'Checked in via QR code'
                }
            )
            
            if not created:
                # Update existing record
                attendance.status = 'present'
                attendance.check_in_time = timezone.now()
                attendance.is_late = False
                attendance.notes = 'Checked in via QR code'
                attendance.save()
            
            # Get session info for response
            session = AttendanceSession.objects.select_related('class_obj').get(id=session_id)
            
            return Response({
                'success': True,
                'message': 'Successfully checked in',
                'attendance_id': attendance.id,
                'session_info': {
                    'id': session.id,
                    'name': session.session_name,
                    'class_name': session.class_obj.class_name,
                    'start_time': session.start_time,
                    'end_time': session.end_time,
                    'location': session.location
                },
                'check_in_time': attendance.check_in_time,
                'is_late': attendance.is_late
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Failed to create attendance record: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'QR validation failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_qr_status(request, session_id):
    """
    Get QR code status for a session
    GET /api/attendance/qr-status/{session_id}/
    """
    try:
        from .qr_service import QRCodeService
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Get QR status
        qr_status = QRCodeService.get_qr_status(session_id, request.user.id)
        
        return Response({
            'success': True,
            'qr_status': qr_status,
            'session_info': {
                'id': session.id,
                'name': session.session_name,
                'class_name': session.class_obj.class_name,
                'is_active': session.is_active,
                'start_time': session.start_time,
                'end_time': session.end_time
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to get QR status: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def revoke_qr_code(request, session_id):
    """
    Revoke QR code for a session
    POST /api/attendance/revoke-qr/{session_id}/
    """
    try:
        from .qr_service import QRCodeService
        
        # Verify session exists and user has permission
        try:
            session = AttendanceSession.objects.get(id=session_id)
            
            # Check if user is the teacher of this session's class
            if session.class_obj.teacher != request.user:
                return Response({
                    'success': False,
                    'error': 'Permission denied'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Revoke QR code
        success = QRCodeService.revoke_qr_token(session_id, request.user.id)
        
        if success:
            return Response({
                'success': True,
                'message': 'QR code revoked successfully'
            })
        else:
            return Response({
                'success': False,
                'error': 'Failed to revoke QR code'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'Failed to revoke QR code: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# WebSocket Integration for Real-time Updates
def send_attendance_update(session_id, attendance_data):
    """
    Send real-time attendance update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'attendance_update',
                'data': {
                    'session_id': session_id,
                    'attendance': attendance_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        # Send to teacher's room
        if attendance_data.get('teacher_id'):
            async_to_sync(channel_layer.group_send)(
                f"user_{attendance_data['teacher_id']}",
                {
                    'type': 'attendance_update',
                    'data': {
                        'session_id': session_id,
                        'attendance': attendance_data,
                        'timestamp': timezone.now().isoformat()
                    }
                }
            )
        
        logger.info(f"Sent attendance update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send attendance update: {e}")


def send_session_status_update(session_id, status_data):
    """
    Send real-time session status update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'session_status_update',
                'data': {
                    'session_id': session_id,
                    'status': status_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent session status update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send session status update: {e}")


def send_qr_code_update(session_id, qr_data):
    """
    Send real-time QR code update via WebSocket
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        channel_layer = get_channel_layer()
        
        # Send to session-specific room
        async_to_sync(channel_layer.group_send)(
            f"session_{session_id}",
            {
                'type': 'qr_code_update',
                'data': {
                    'session_id': session_id,
                    'qr_code': qr_data,
                    'timestamp': timezone.now().isoformat()
                }
            }
        )
        
        logger.info(f"Sent QR code update for session {session_id}")
        
    except Exception as e:
        logger.error(f"Failed to send QR code update: {e}")