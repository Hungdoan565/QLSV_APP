from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from .models import Student
from .serializers import StudentSerializer, StudentCreateSerializer


class StudentListCreateView(generics.ListCreateAPIView):
    """List and create students"""
    queryset = Student.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return StudentCreateSerializer
        return StudentSerializer
    
    def get_queryset(self):
        queryset = Student.objects.all()
        search = self.request.query_params.get('search', None)
        if search is not None:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(student_id__icontains=search) |
                Q(email__icontains=search)
            )
        return queryset


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete a student"""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_create_students(request):
    """Bulk create students (simplified version without Excel import)"""
    try:
        students_data = request.data.get('students', [])
        created_students = []
        errors = []
        
        for student_data in students_data:
            serializer = StudentCreateSerializer(data=student_data)
            if serializer.is_valid():
                student = serializer.save()
                created_students.append(StudentSerializer(student).data)
            else:
                errors.append(serializer.errors)
        
        return Response({
            'success': True,
            'created_count': len(created_students),
            'created_students': created_students,
            'errors': errors
        })
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_excel(request):
    """Import students from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'message': 'Only Excel files (.xlsx, .xls) are supported'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Parse Excel file with openpyxl
        from openpyxl import load_workbook
        import io
        
        # Load workbook from uploaded file
        workbook = load_workbook(io.BytesIO(excel_file.read()))
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value.lower().replace(' ', '_') if cell.value else '')
        
        # Expected columns mapping
        column_mapping = {
            'student_id': ['student_id', 'mssv', 'id'],
            'first_name': ['first_name', 'ten', 'ho_ten', 'name'],
            'last_name': ['last_name', 'ho', 'surname'],
            'email': ['email', 'mail'],
            'phone': ['phone', 'sdt', 'telephone'],
            'gender': ['gender', 'gioi_tinh', 'sex'],
            'date_of_birth': ['date_of_birth', 'ngay_sinh', 'birthday'],
            'address': ['address', 'dia_chi', 'location']
        }
        
        # Map headers to expected fields
        field_mapping = {}
        for field, possible_names in column_mapping.items():
            for header in headers:
                if header in possible_names:
                    field_mapping[field] = headers.index(header)
                    break
        
        created_students = []
        errors = []
        
        # Process each row
        for row_num in range(2, worksheet.max_row + 1):
            try:
                row_data = {}
                for field, col_index in field_mapping.items():
                    cell_value = worksheet.cell(row=row_num, column=col_index + 1).value
                    row_data[field] = str(cell_value).strip() if cell_value else ''
                
                # Set defaults for missing fields
                if not row_data.get('gender'):
                    row_data['gender'] = 'male'
                if not row_data.get('date_of_birth'):
                    row_data['date_of_birth'] = '2000-01-01'
                
                # Validate required fields
                if not row_data.get('student_id') or not row_data.get('first_name'):
                    errors.append({
                        'row': row_num,
                        'error': 'Missing required fields: student_id and first_name',
                        'data': row_data
                    })
                    continue
                
                # Check if student already exists
                if Student.objects.filter(student_id=row_data['student_id']).exists():
                    errors.append({
                        'row': row_num,
                        'error': f'Student with ID {row_data["student_id"]} already exists',
                        'data': row_data
                    })
                    continue
                
                # Create student
                serializer = StudentCreateSerializer(data=row_data)
                if serializer.is_valid():
                    student = serializer.save()
                    created_students.append(StudentSerializer(student).data)
                else:
                    errors.append({
                        'row': row_num,
                        'errors': serializer.errors,
                        'data': row_data
                    })
                    
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'error': str(e),
                    'data': f'Row {row_num} processing failed'
                })
        
        return Response({
            'success': True,
            'message': f'Successfully imported {len(created_students)} students from Excel file',
            'created_count': len(created_students),
            'created_students': created_students,
            'errors': errors,
            'details': {
                'total_rows_processed': worksheet.max_row - 1,
                'successful_imports': len(created_students),
                'failed_imports': len(errors)
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_excel(request):
    """Export students to Excel/CSV file"""
    try:
        import csv
        import io
        
        # Get all students
        students = Student.objects.all()
        
        # Create CSV content
        output = io.StringIO()
        fieldnames = [
            'student_id', 'first_name', 'last_name', 'email', 
            'phone', 'gender', 'date_of_birth', 'address', 'is_active'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for student in students:
            writer.writerow({
                'student_id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'email': student.email,
                'phone': student.phone or '',
                'gender': student.gender,
                'date_of_birth': student.date_of_birth,
                'address': student.address or '',
                'is_active': student.is_active
            })
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        return response
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_statistics(request):
    """Get comprehensive student statistics"""
    try:
        from django.db.models import Count, Q
        from datetime import date, timedelta
        
        # Basic counts
        total_students = Student.objects.count()
        active_students = Student.objects.filter(is_active=True).count()
        inactive_students = Student.objects.filter(is_active=False).count()
        
        # Gender distribution
        gender_stats = Student.objects.values('gender').annotate(count=Count('id'))
        
        # Age distribution
        today = date.today()
        age_groups = {
            '18-20': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=20*365),
                date_of_birth__lt=today - timedelta(days=18*365)
            ).count(),
            '21-25': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=25*365),
                date_of_birth__lt=today - timedelta(days=21*365)
            ).count(),
            '26-30': Student.objects.filter(
                date_of_birth__gte=today - timedelta(days=30*365),
                date_of_birth__lt=today - timedelta(days=26*365)
            ).count(),
            '30+': Student.objects.filter(
                date_of_birth__lt=today - timedelta(days=30*365)
            ).count(),
        }
        
        # Recent registrations (last 30 days)
        thirty_days_ago = today - timedelta(days=30)
        recent_registrations = Student.objects.filter(
            created_at__gte=thirty_days_ago
        ).count()
        
        # Students with missing information
        missing_phone = Student.objects.filter(phone__isnull=True).count()
        missing_address = Student.objects.filter(address__isnull=True).count()
        
        return Response({
            'total_students': total_students,
            'active_students': active_students,
            'inactive_students': inactive_students,
            'gender_distribution': list(gender_stats),
            'age_groups': age_groups,
            'recent_registrations': recent_registrations,
            'missing_information': {
                'missing_phone': missing_phone,
                'missing_address': missing_address
            },
            'completion_rate': round(
                ((total_students - missing_phone - missing_address) / total_students * 100) 
                if total_students > 0 else 0, 2
            )
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
